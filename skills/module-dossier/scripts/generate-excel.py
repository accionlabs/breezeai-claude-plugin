#!/usr/bin/env python3
"""
Generate the capability-matrix Excel spreadsheet from aggregated module data.

Usage:
    python3 generate-excel.py <excel-data.json> <output.xlsx>

Input JSON schema:
    {
      "rows": [
        {
          "domain":      "Customer Account Management",
          "module_num":  "1",
          "module_name": "Profile Management",
          "item_num":    "1.1",
          "capability":  "Manage Primary Sponsor",
          "personas":    ["Case Manager", "Enrollment Specialist"],
          "files":       ["onecms/CMS/Controls/Sponsor/ManageSponsor.ascx", ...],
          "procs":       ["STD_GetSponsor", "STD_SaveSponsor"],
          "tables":      ["std.Sponsor", "std.SponsorAddress"]
        }
      ]
    }

Output:
    Single-sheet .xlsx with 13 columns, one row per item.
    Files classified into Client / Service / Façade / Data Access / Jobs by path prefix.
"""

import json
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Layer classification — case-insensitive substring on normalised path
# ---------------------------------------------------------------------------

LAYERS = [
    ("client", [
        "onecms/cms",
        "onecms/presenters/kucare.presenters",
        "onecms/subsidy",
    ]),
    ("service", [
        "onecms/easydraftservice",
        "onecms/framework/kucare.services",
        "onecms/framework/kucare.enterpriseservices",
        "onecms/framework/kucare.enterpriseextservices",
        "onecms/framework/kucare.enrollmentrefactor/domain",
    ]),
    ("facade", [
        "onecms/framework/kucare.facade",
        "onecms/framework/kucare.enrollmentrefactor/facade",
    ]),
    ("dal", [
        "onecms/framework/kucare.repositories",
        "onecms/framework/kucare.data",
        "onecms/framework/kucare.enrollmentrefactor/repositories",
        "onecms/dataaccessmanagement",
    ]),
    ("jobs", [
        "onecms/kucare.windowservices",
    ]),
]


def classify(path):
    norm = path.replace("\\", "/").lower()
    for layer, prefixes in LAYERS:
        if any(p in norm for p in prefixes):
            return layer
    return None


def split_files(raw_files):
    buckets: dict[str, list[str]] = {"client": [], "service": [], "facade": [], "dal": [], "jobs": []}
    dropped = 0
    for f in raw_files:
        layer = classify(f)
        if layer:
            norm_f = f.replace("\\", "/")
            # strip leading "onecms/" prefix used only for classification matching
            display = norm_f[len("onecms/"):] if norm_f.lower().startswith("onecms/") else norm_f
            if display not in buckets[layer]:
                buckets[layer].append(display)
        else:
            dropped += 1
    return buckets, dropped


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

HEADERS = [
    "Domain",
    "Module #",
    "Module Name",
    "Item #",
    "Capability",
    "Personas",
    "List of Source Code Files (Client)",
    "List of Source Code Files (Service Layer)",
    "List of Source Code Files (Façade Layer)",
    "List of Source Code Files (Data Access Layer)",
    "List of Stored Procedures",
    "List of Tables",
    "Jobs",
]

COL_WIDTHS = [22, 10, 24, 8, 36, 24, 48, 48, 48, 48, 36, 30, 40]

DARK_BLUE  = "1F4E79"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"

HDR_FONT   = Font(name="Calibri", color=WHITE, bold=True, size=11)
HDR_FILL   = PatternFill("solid", fgColor=DARK_BLUE)
HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT  = Font(name="Calibri", size=10)
ALT_FILL   = PatternFill("solid", fgColor=LIGHT_BLUE)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

THIN = Side(style="thin", color="CCCCCC")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_excel(rows: list[dict], output_path: str) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capability Matrix"

    # Header
    for ci, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font   = HDR_FONT
        cell.fill   = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BOX
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    total_dropped = 0

    for ri, row in enumerate(rows, start=2):
        buckets, dropped = split_files(row.get("files", []))
        total_dropped += dropped

        personas = row.get("personas", [])
        personas_str = "\n".join(personas) if isinstance(personas, list) else str(personas)

        values = [
            row.get("domain", ""),
            row.get("module_num", ""),
            row.get("module_name", ""),
            row.get("item_num", ""),
            row.get("capability", ""),
            personas_str,
            "\n".join(sorted(buckets["client"])),
            "\n".join(sorted(buckets["service"])),
            "\n".join(sorted(buckets["facade"])),
            "\n".join(sorted(buckets["dal"])),
            "\n".join(sorted(row.get("procs", []))),
            "\n".join(sorted(row.get("tables", []))),
            "\n".join(sorted(buckets["jobs"])),
        ]

        use_alt = (ri % 2 == 0)
        for ci, value in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font      = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.border    = BOX
            if use_alt:
                cell.fill = ALT_FILL

        # Auto row height hint (openpyxl doesn't auto-size; set a generous default)
        file_lines = max(
            len(buckets["client"]), len(buckets["service"]),
            len(buckets["facade"]), len(buckets["dal"]),
            len(buckets["jobs"]),
            len(row.get("procs", [])), len(row.get("tables", [])),
            1,
        )
        ws.row_dimensions[ri].height = max(16, min(file_lines * 15, 200))

    wb.save(output_path)
    return total_dropped


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 3:
        print(__doc__, file=sys.stderr)
        sys.exit(1)

    data_file   = sys.argv[1]
    output_file = sys.argv[2]

    with open(data_file, "r", encoding="utf-8") as f:
        payload = json.load(f)

    rows = payload.get("rows", [])
    if not rows:
        print("No rows found in input JSON.", file=sys.stderr)
        sys.exit(1)

    dropped = write_excel(rows, output_file)
    print(
        f"Written to {output_file} — "
        f"{len(rows)} rows"
        + (f", {dropped} unclassified files dropped" if dropped else "")
    )


if __name__ == "__main__":
    main()
